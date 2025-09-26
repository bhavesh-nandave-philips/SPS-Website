from flask import Flask, render_template_string, request, send_file
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
import io
import os
import re
import requests
from zipfile import ZipFile
from urllib.parse import urljoin
import tempfile
import traceback

app = Flask(__name__)

WEBSITE_URL = "https://nlyehvdcs3vw102.code1.emi.philips.com/SPS/Identification_Database/#"

HTML_TEMPLATE = '''
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>SPS Image Check</title>
  <style>
    body {
      margin: 0;
      font-family: Arial, sans-serif;
      background-color: #87ceeb; /* sky blue */
      height: 100vh;
      display: flex;
      flex-direction: column;
      justify-content: center;
      align-items: center;
      color: #003366;
    }
    h1 {
      font-size: 3rem;
      margin-bottom: 2rem;
      font-weight: bold;
      text-align: center;
      text-shadow: 1px 1px 3px rgba(0,0,0,0.3);
    }
    form {
      background: white;
      padding: 2rem 3rem;
      border-radius: 8px;
      box-shadow: 0 4px 10px rgba(0,0,0,0.2);
      display: flex;
      flex-direction: column;
      align-items: center;
      min-width: 320px;
    }
    input[type="file"] {
      margin-bottom: 1.5rem;
      font-size: 1rem;
    }
    button {
      background-color: #003366;
      color: white;
      border: none;
      padding: 0.75rem 2rem;
      font-size: 1.1rem;
      font-weight: bold;
      border-radius: 5px;
      cursor: pointer;
      transition: background-color 0.3s ease;
    }
    button:hover {
      background-color: #002244;
    }
    p.message {
      margin-top: 1.5rem;
      font-size: 1.1rem;
      color: #222;
      font-weight: 600;
      text-align: center;
    }
  </style>
</head>
<body>
  <h1>SPS Image Check</h1>
  <form method="post" action="/process" enctype="multipart/form-data" novalidate>
      <input type="file" name="file" accept=".xlsx" required>
      <button type="submit">Start</button>
  </form>

  {% if message %}
    <p class="message">{{ message }}</p>
  {% endif %}
</body>
</html>
'''

excluded_substrings = [
    "Images/shield.png",
    "Images\\shield.png",
    "cleardot.gif",
    "google.com/images/",
    "translate.googleapis.com/",
    "translate/",
    "translate.",
    "translate_",
    "logo.png"
]

def sanitize_filename(text):
    return re.sub(r'[^a-zA-Z0-9_\-]', '_', text)

def download_images_for_serial(page, serial, download_root):
    image_paths = []
    try:
        print(f"Navigating to main URL for serial {serial}...")
        page.goto(WEBSITE_URL)
        page.wait_for_load_state("networkidle")

        # Search serial
        page.fill('#searchValue', serial)
        page.click('#search')
        page.wait_for_load_state("networkidle")

        # Scroll to bottom to load images
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        page.wait_for_timeout(1000)

        img_elems = page.query_selector_all("img")

        safe_serial = "".join(c for c in serial if c.isalnum() or c in ('-', '_'))

        for i, img in enumerate(img_elems, start=1):
            src = img.get_attribute("src")
            if not src or any(excl in src for excl in excluded_substrings):
                continue

            source_text = "unknown_source"
            try:
                parent_a = img.query_selector("xpath=ancestor::a[1]")
                span = None
                if parent_a:
                    span = parent_a.query_selector("xpath=following-sibling::span[1]")
                if not span:
                    span = img.query_selector("xpath=following-sibling::span[1]")
                if span:
                    source_raw = span.inner_text().strip()
                    if source_raw.lower().startswith("source:"):
                        source_text = source_raw[7:].strip()
                    else:
                        source_text = source_raw
                    source_text = re.sub(r'[\s\(\)\.]+', '_', source_text).strip('_').lower()
            except Exception:
                # Just ignore if source extraction fails
                pass

            full_url = urljoin(WEBSITE_URL, src)
            try:
                resp = requests.get(full_url, timeout=10)
                resp.raise_for_status()
            except Exception as e:
                print(f"Failed to fetch image: {full_url} error: {e}")
                continue

            ext = os.path.splitext(full_url)[1].lower()
            if ext not in [".jpg", ".jpeg", ".png", ".gif", ".bmp", ".svg", ".webp"]:
                ext = ".jpg"

            filename = f"{safe_serial}_{i}_{sanitize_filename(source_text)}{ext}"
            fpath = os.path.join(download_root, filename)

            with open(fpath, "wb") as f:
                f.write(resp.content)
            image_paths.append(fpath)

    except Exception as e:
        print(f"Error processing serial {serial}: {e}")
        traceback.print_exc()

    return image_paths

@app.route('/', methods=["GET"])
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/process', methods=["POST"])
def process():
    uploaded = request.files.get("file")
    if not uploaded:
        return render_template_string(HTML_TEMPLATE, message="No file uploaded")

    try:
        wb = load_workbook(uploaded, data_only=True)
        ws = wb.active
    except Exception as e:
        return render_template_string(HTML_TEMPLATE, message=f"Cannot read Excel: {e}")

    serials = [str(row[0]).strip() for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]]
    if not serials:
        return render_template_string(HTML_TEMPLATE, message="No serials found in the file")

    results = {}

    try:
        with sync_playwright() as p, tempfile.TemporaryDirectory() as tmpdir:
            print("Launching browser...")
            browser = p.chromium.launch(headless=True, args=["--no-sandbox", "--disable-setuid-sandbox"])
            page = browser.new_page()

            for serial in serials:
                print(f"Processing serial: {serial}")
                image_paths = download_images_for_serial(page, serial, tmpdir)
                results[serial] = "YES" if image_paths else "NO"

            browser.close()

            # Update Excel with results
            ws.cell(row=1, column=2, value="ImagesFound")
            for idx, serial in enumerate(serials, start=2):
                ws.cell(row=idx, column=2, value=results.get(serial, "NO"))

            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            zip_buffer = io.BytesIO()
            with ZipFile(zip_buffer, "w") as zipf:
                for fname in os.listdir(tmpdir):
                    fpath = os.path.join(tmpdir, fname)
                    zipf.write(fpath, arcname=fname)
                zipf.writestr("results.xlsx", excel_buffer.getvalue())

            zip_buffer.seek(0)

            return send_file(
                zip_buffer,
                as_attachment=True,
                download_name="sps_images_and_results.zip",
                mimetype="application/zip"
            )
    except Exception as e:
        print(f"Error during processing: {e}")
        traceback.print_exc()
        return render_template_string(HTML_TEMPLATE, message=f"Processing error: {e}")

if __name__ == "__main__":
    # Make sure to run in your terminal before running this script:
    # > playwright install
    import os
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
